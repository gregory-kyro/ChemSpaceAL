# Necessary import statements
import torch
import torch.nn.functional as F
import rdkit
from rdkit import Chem
import rdkit.Chem.Descriptors
import pandas as pd
import re
from openpyxl import load_workbook
import tqdm
import pickle
from sklearn.cluster import KMeans

from Configuration import *
from GPT_Dataset import *
from GPT_Model import *

@torch.no_grad()
def sample(model, x, steps, temperature=1.0):
    """
    Sample a sequence of tokens from a given model
    """
    block_size = model.get_block_size()
    model.eval()
    for k in range(steps):
        x_cond = x if x.size(1) <= block_size else x[:, -block_size:] # Get last block_size elements
        logits, _ = model(x_cond) # Get logits from model
        logits = logits[:, -1, :] / temperature # Extract logits for next token
        probs = F.softmax(logits, dim=-1)
        ix = torch.multinomial(probs, 1)
        x = torch.cat((x, ix), dim=1) # Concatenate token with sequence
    return x


# Function to check the novelty of generated molecules relative to a training list
def check_novelty(generated, train_list, sig_digs=3, multiplier=100, denominator=None, subtracted=True, show_work=False):
    total_train = set()
    for train in train_list:
        total_train = total_train | train
    repeated = generated & total_train
    if denominator is None:
        denominator = len(generated)
    if subtracted:
        if show_work:
            out = np.round(multiplier*(1-len(repeated)/denominator), sig_digs)
            return f"{multiplier}*(1-{len(repeated)}/{denominator}) = {out}"
        else:
            return np.round(multiplier*(1-len(repeated)/denominator), sig_digs)
    else:
        if show_work:
            out = np.round(multiplier*len(repeated)/denominator, sig_digs)
            return f"{multiplier} * {len(repeated)}/{denominator} = {out}"
        else:
            return np.round(multiplier*len(repeated)/denominator, sig_digs)


def canonic_smiles(smiles_or_mol):
    """
    Convert SMILES or RDKit mol object to canonical SMILES string
    """
    mol = get_mol(smiles_or_mol) # Convert to RDKit mol object
    if mol is None:
        return None
    return Chem.MolToSmiles(mol) # Convert to canonical SMILES


def load_data(config_dict, mode='Active Learning', forced_block_size=None):
    """
    Load data to be used for either pretraining or active learning
    """

    # Pretraining
    if mode == 'Pretraining':
        # Get compression
        if 'gz' in config_dict["train_path"]:
            compression = 'gzip'
        else:
            compression = None

        # Slice data if set in configuration dictionary
        if (cut:=config_dict["slice_data"]):
            train_data = pd.read_csv(config_dict["train_path"], compression=compression)[:cut]
            val_data = pd.read_csv(config_dict["val_path"], compression=compression)[:cut]

        else:
            train_data = pd.read_csv(config_dict["train_path"], compression=compression)
            val_data = pd.read_csv(config_dict["val_path"], compression=compression)
        iterators = (train_data[config_dict['smiles_key']].values, val_data[config_dict['smiles_key']].values)
        assert len(train_data) == len(train_data[config_dict['smiles_key']].values), "There's no reason why this shouldn't be true"
    
    elif mode == 'Active Learning':
        print(f"Loading Active Learning dataset from", '/'.join(config_dict["al_path"].split('/')[6:]))
        al_data = pd.read_csv(config_dict["al_path"])
        iterators = (al_data[config_dict['smiles_key']].values, )
    else:
        raise KeyError(f"Only pretraining and active learning are currently supported")

    # Compile pattern into a regular expression object that can be used for matching operations
    regex = re.compile(Config.REGEX_PATTERN)
    char_set = {'<', '!', '~'}

    max_len = 0
    for iterator in iterators:
        for i in iterator:
            chars = regex.findall(i.strip())
            max_len = max(max_len, len(chars))
            for char in chars:
                char_set.add(char)

    chars = sorted(list(char_set))
    max_len += 1    # Accounting for the start token, which hasn't been added yet
    if forced_block_size is not None:
        assert mode == 'Active Learning', "Cannot force a block size in pretraining"
        max_len = forced_block_size

    datasets = []
    for iterator in iterators:
        padded = ['!' + i + '~' + '<'*(max_len - 1 - len(regex.findall(i.strip()))) for i in iterator]
        dataset = SMILESDataset(data=padded, chars=chars, block_size=max_len, len_data=len(iterator))
        datasets.append(dataset)
        
    dataset.export_desc_attributes(config_dict["desc_path"])
    
    return datasets


# Function to export given metrics to an Excel workbook
def export_metrics_to_workbook(metrics, fname):
    # Map each metric to its corresponding column in the Excel sheet
    metric_to_col = {
        'generated': 'B',
        'valid': 'C',
        'unique': 'D',
        'validity': 'E',
        '% unique (rel. to generated)': 'F',
        '% unique (rel. to valid)': 'G',
        '% novelty (rel. to train set)': 'H',
        '% novelty (rel. to train+AL sets)': 'I',
        '% repetitions (from AL0 training set)': 'J',
        '% repetitions (from AL1 training set)': 'K',
        '% repetitions (from AL2 training set)': 'L',
        '% repetitions (from AL3 training set)': 'M',
        '% repetitions (from AL4 training set)': 'N',
        '% repetitions (from AL5 training set)': 'O',
        '% repetitions (from AL6 training set)': 'P',
        '% repetitions (from AL7 training set)': 'Q',
        '% repetitions (from scored from round 0)': 'R',
        '% repetitions (from scored from round 1)': 'S',
        '% repetitions (from scored from round 2)': 'T',
        '% repetitions (from scored from round 3)': 'U',
        '% repetitions (from scored from round 4)': 'V',
        '% repetitions (from scored from round 5)': 'W',
        '% repetitions (from scored from round 6)': 'X',
        '% repetitions (from scored from round 7)': 'Y',
        '% fraction of AL0 training set in generated': 'Z',
        '% fraction of AL1 training set in generated': 'AA',
        '% fraction of AL2 training set in generated': 'AB',
        '% fraction of AL3 training set in generated': 'AC',
        '% fraction of AL4 training set in generated': 'AD',
        '% fraction of AL5 training set in generated': 'AE',
        '% fraction of AL6 training set in generated': 'AF',
        '% fraction of AL7 training set in generated': 'AG',
        '% fraction of scored from round 0 in generated': 'AH',
        '% fraction of scored from round 1 in generated': 'AI',
        '% fraction of scored from round 2 in generated': 'AJ',
        '% fraction of scored from round 3 in generated': 'AK',
        '% fraction of scored from round 4 in generated': 'AL',
        '% fraction of scored from round 5 in generated': 'AM',
        '% fraction of scored from round 6 in generated': 'AN',
        '% fraction of scored from round 7 in generated': 'AO'}
    
    # Use the openpyxl library to load an Excel workbook
    wb = load_workbook(filename=f"{BASE_PATH}Generative_ML Logbook.xlsx")

    # Loop through specific sheets and find the first nonempty row to append metrics
    for i, sheet in enumerate(('generated_logbook_abs', 'generated_logbook_rel')):
        ws = wb[sheet]
        row = 3
        while ws[f'A{row}'].value is not None:
            row += 1
        ws[f'A{row}'] = fname
        for metric, value in metrics.items():
            # Processing the value if it contains an equation
            if isinstance(value, str) and '=' in value:
                if i == 0:
                    value = value.split(' = ')[0].split(' * ')[1]
                else:
                    value = value.split(' = ')[1]
            ws[f'{metric_to_col[metric]}{row}'] = value
    # Save the updated workbook
    wb.save(filename=f"{Config.BASE_PATH}Generative_ML Logbook.xlsx")


# Function to convert a SMILES string into an RDKit Mol object
def get_mol(smile_string):
    mol = Chem.MolFromSmiles(smile_string)
    if mol is None:
        return None
    try:
        Chem.SanitizeMol(mol)
    except ValueError:
        return None
    return mol


# Function to write a dictionary to a text file
def dump_dic_to_text(dic, path, header=None):
    with open(path, 'w') as f:
        if header is not None:
            f.write(f"{header}\n")
        for key, value in dic.items():
            f.write(f'{key}: {value}\n')


# Function to generate SMILES strings for molecules using a given model and parameters
def generate_SMILES(config_dict):
    regex = re.compile(Config.REGEX_PATTERN)
    dataset = SMILESDataset()
    dataset.load_desc_attributes(config_dict["desc_path"])

    mconf = GPTConfig(dataset.vocab_size, dataset.block_size, **config_dict)
    model = GPT(mconf).to(config_dict["device"])
    model.load_state_dict(torch.load(config_dict["load_ckpt_path"], map_location=torch.device(config_dict["device"])))
    model.to(config_dict["device"])
    torch.compile(model)

    # load parameters into the model
    block_size = model.get_block_size()
    assert (block_size == dataset.block_size), "Warning: model block size and dataset block size are different"
    molecules_list, molecules_set = [], set()
    completions = []
    pbar = tqdm.tqdm()
    while True:
        pbar.update()
        pbar.set_description(f"generated {len(molecules_set)} unique molecules")
        # create an input tensor by converting 'context' to a tensor of token indices, repeat this batch times along the batch dimension
        x = (torch.tensor([dataset.stoi[s] for s in regex.findall(config_dict["generation_context"])], dtype=torch.long,)[None, ...]
            .repeat(config_dict["gen_batch_size"], 1).to(config_dict["device"]))
        y = sample(model, x, block_size, temperature=config_dict["inference_temp"])
        for gen_mol in y:
            completion = "".join([dataset.itos[int(i)] for i in gen_mol])  # convert generated molecule from list of integers to list of strings and concatenate to one string
            completions.append(completion)
            if completion[0] == '!' and completion[1] == '~':
                completion = '!' + completion[2:]
            if "~" not in completion: continue
            mol_string = completion[1 : completion.index("~")]
            mol = get_mol(mol_string)  # convert the string representation of the molecule to an rdkit Mol object
            if mol is not None:
                molecules_list.append(Chem.MolToSmiles(mol))
                molecules_set.add(Chem.MolToSmiles(mol))
        if len(molecules_set) >= config_dict["gen_size"]:
            break
    pbar.close()

    completions_df = pd.DataFrame({"smiles": completions})
    completions_df.to_csv(config_dict["generation_path"]+ f"_temp{config_dict['inference_temp']}_completions.csv")
    molecules_df = pd.DataFrame({"smiles": list(molecules_set)})
    molecules_df.to_csv(config_dict["generation_path"]+ f"_temp{config_dict['inference_temp']}_processed.csv")
    characterize_generated_molecules(config_dict, molecules_list)


# Function to characterize the generated molecules using various metrics
def characterize_generated_molecules(config_dict, molecules_list=None):
    # Load completions and process the generated molecules
    completions = pd.read_csv(config_dict["generation_path"]+ f"_temp{config_dict['inference_temp']}_completions.csv")['smiles']
    molecules_set = set(pd.read_csv(config_dict["generation_path"]+ f"_temp{config_dict['inference_temp']}_processed.csv")['smiles'])
    if molecules_list is None:
        molecules_list = []
        for completion in tqdm(completions, total=len(completions)):
            if completion[0] == '!' and completion[1] == '~':
                completion = '!' + completion[2:]
            if "~" not in completion: continue
            mol_string = completion[1 : completion.index("~")]
            mol = get_mol(mol_string)  # convert the string representation of the molecule to an rdkit Mol object
            if mol is not None:
                molecules_list.append(Chem.MolToSmiles(mol))

    assert molecules_set == set(molecules_list), "Warning: set(molecules_list) and molecules_set are different"
    train_data = set(pd.read_csv(config_dict["train_path"])[config_dict["smiles_key"]])
    scored_sets = {i: set(pd.read_csv(path)['smiles']) for i, path in enumerate(config_dict['diffdock_scored_path_list'])}
    al_sets = {i: set(pd.read_csv(path)['smiles']) for i, path in enumerate(config_dict['al_trainsets_path_list'])}

    # Calculate various metrics related to the generated molecules
    multiplier = 100
    metrics = {
        "generated": len(completions), "valid": len(molecules_list), "unique": len(molecules_set),
        "validity": np.round(multiplier*len(molecules_list)/len(completions), 3),
        "% unique (rel. to generated)": np.round(multiplier*len(molecules_set)/len(completions), 3),
        "% unique (rel. to valid)": np.round(multiplier*len(molecules_set)/len(molecules_list), 3),
        "% novelty (rel. to train set)": check_novelty(molecules_set, (train_data,), multiplier=multiplier),
        "% novelty (rel. to train+AL sets)": check_novelty(molecules_set, (train_data, *list(al_sets.values())), multiplier=multiplier),
    }
    for al_round, al_set in al_sets.items():
        metrics[f"% repetitions (from AL{al_round} training set)"] = check_novelty(molecules_set, (al_set,), subtracted=False, multiplier=multiplier, show_work=True)
    for score_round, score_set in scored_sets.items():
        metrics[f"% repetitions (from scored from round {score_round})"] = check_novelty(molecules_set, (score_set,), subtracted=False, multiplier=multiplier, show_work=True)
    for al_round, al_set in al_sets.items():
        metrics[f"% fraction of AL{al_round} training set in generated"] = check_novelty(molecules_set, (al_set,), subtracted=False, multiplier=multiplier, denominator=len(al_set), show_work=True)
    for score_round, score_set in scored_sets.items():
        metrics[f"% fraction of scored from round {score_round} in generated"] = check_novelty(molecules_set, (score_set,), subtracted=False, multiplier=multiplier, denominator=len(score_set), show_work=True)

    # Dump the metrics to a text file and export them to a workbook
    dump_dic_to_text(metrics, config_dict["generation_path"]+ f"_temp{config_dict['inference_temp']}_metrics.txt")
    export_metrics_to_workbook(metrics, config_dict["generation_path"].split('/')[-1])


def descriptors_for_gpt_predictions(config_dict):
    # Read the SMILES data from the CSV file specified in the config
    gpt_mols = pd.read_csv(config_dict['path_to_predicted'])
    smiles_set = set(gpt_mols['smiles'].to_list())

    # Iterate through each file path in the scored path list
    for scored_file in config_dict['diffdock_scored_path_list']:
        for smile in pd.read_csv(scored_file)['smiles'].values:
            # Ensure that each smile is a string
            assert isinstance(smile, str), f"{smile} is not a string"
            smiles_set.add(smile)

    keySet = None
    keyToData = {}

    # Create a progress bar to monitor progress through the smiles_set
    pbar = tqdm.tqdm(smiles_set, total=len(smiles_set))

    # Process each smile in the set
    for smile in pbar:
        mol = rdkit.Chem.MolFromSmiles(smile)
        if not mol:
            continue

        # Calculate the molecular descriptors using RDKit
        mol_data = rdkit.Chem.Descriptors.CalcMolDescriptors(mol)

        # If keySet is None, initialize it with the keys from mol_data
        if keySet is None:
            keySet = set(mol_data.keys())

        # Add the calculated descriptors to keyToData
        for key in keySet:
            keyToData.setdefault(key, []).append(mol_data[key])

        # Add the original smile to keyToData
        keyToData.setdefault('smiles', []).append(smile)

    # Convert the collected data to a DataFrame
    gpt_df = pd.DataFrame(keyToData)

    # Save the DataFrame to a pickle file for later use
    gpt_df.to_pickle(config_dict['path_to_descriptors'])

    return gpt_df
    

def project_into_pca_space(config):
    # Load the scaler and PCA objects from the specified file
    scaler, pca = pickle.load(open(config["path_to_pca"], 'rb'))

    # Read the descriptors from the pickled DataFrame file
    gptMols = pd.read_pickle(config["path_to_descriptors"])

    # First, apply the scaler to standardize the descriptors
    scaled_data = scaler.transform(gptMols[scaler.get_feature_names_out()])

    # Next, apply the PCA transformation to project the scaled data into PCA space
    pca_data = pca.transform(scaled_data)

    return pca_data


def _cluster_mols_experimental_loss(mols, n_clusters, n_iter):
    # Iterate to find the KMeans clustering with the minimum inertia (loss)
    min_loss, best_kmeans = float('inf'), None
    for _ in range(n_iter):
        kmeans = KMeans(n_clusters=n_clusters, n_init='auto', init='k-means++').fit(mols)
        if kmeans.inertia_ < min_loss:
            min_loss = kmeans.inertia_
            best_kmeans = kmeans
    return best_kmeans


def _cluster_mols_experimental_variance(mols, n_clusters, n_iter):
    # Iterate to find the KMeans clustering with the maximum variance
    max_variance, best_kmeans = float('-inf'), None
    for _ in range(n_iter):
        kmeans = KMeans(n_clusters=n_clusters, n_init='auto', init='k-means++').fit(mols)
        counts = np.unique(kmeans.labels_, return_counts=True)[1]
        if (variance:=np.var(counts)) > max_variance:
            max_variance = variance
            best_kmeans = kmeans
    return best_kmeans


def _cluster_mols_experimental_mixed(mols, n_clusters, n_iter, mixed_objective_loss_quantile):
    # Mixed approach, combining inertia and variance objectives
    inertias, variances, km_objs = [], [], []
    for _ in range(n_iter):
        kmeans = KMeans(n_clusters=n_clusters, n_init='auto', init='k-means++').fit(mols)
        inertias.append(kmeans.inertia_)
        counts = np.unique(kmeans.labels_, return_counts=True)[1]
        variances.append(np.var(counts))
        km_objs.append(kmeans)
    loss_var_kmeans_triples = sorted(zip(inertias, variances, km_objs), key=lambda x: x[0])
    lowest_n = loss_var_kmeans_triples[:int(len(loss_var_kmeans_triples) * mixed_objective_loss_quantile)]
    sorted_by_variance = sorted(lowest_n, key=lambda x: x[1])
    return sorted_by_variance[0][2]


def _cluster_mols_experimental(mols, n_clusters, save_path, n_iter=1, objective='loss', mixed_objective_loss_quantile=0.1):
    # Main clustering function with support for various objectives
    if n_iter == 1:
        kmeans = KMeans(n_clusters=n_clusters, n_init='auto', init='k-means++').fit(mols)
    elif objective == 'loss':
        kmeans = _cluster_mols_experimental_loss(mols, n_clusters, n_iter)
    elif objective == 'variance':
        kmeans = _cluster_mols_experimental_variance(mols, n_clusters, n_iter)
    elif objective == 'mixed':
        kmeans = _cluster_mols_experimental_mixed(mols, n_clusters, n_iter, mixed_objective_loss_quantile)
    else:
        raise ValueError(f'Unknown objective {objective}')

    # Save the kmeans object to a file
    pickle.dump(kmeans, open(save_path, 'wb'))
    return kmeans


def cluster_and_sample(mols, config_dict, ensure_correctness=False, load_kmeans=False):
    
    n_clusters = config_dict['n_clusters']
    n_samples = config_dict['samples_per_cluster']
    path_to_pca = config_dict['path_to_pca']
    
    # Ensure that the requested number of clusters and samples doesn't exceed available molecules
    assert n_clusters * n_samples <= len(mols), f"{n_clusters=} * {n_samples=} = {n_clusters*n_samples} requested but only {len(mols)} molecules provided"

    # If correctness is to be ensured, path_to_pca must be provided
    if ensure_correctness:
        assert path_to_pca is not None, "path_to_pca must be provided to ensure correctness"
        scaler, pca = pickle.load(open(path_to_pca, 'rb'))

    # Load existing KMeans or create a new one
    if load_kmeans:
        kmeans = pickle.load(open(config_dict['kmeans_save_path'], 'rb'))
    else:
        kmeans = _cluster_mols_experimental(mols=mols, n_iter=100, n_clusters=n_clusters, save_path=config_dict["kmeans_save_path"], objective='mixed', mixed_objective_loss_quantile=0.05)

    mols_smiles = pd.read_pickle(config["path_to_descriptors"])['smiles']
    
    assert len(kmeans.labels_) == len(mols_smiles), "Number of labels differs from number of molecules"
    
    scored_smiles = set()
    for scored_file in config_dict['diffdock_scored_path_list']:
        for smile in pd.read_csv(scored_file)['smiles'].values:
            scored_smiles.add(smile)

    # Group molecules by cluster labels
    cluster_to_mols = {}
    for mol, label, smile in zip(mols, kmeans.labels_, mols_smiles):
        if smile in scored_smiles:
            continue
        cluster_to_mols.setdefault(label, []).append(smile)
        
        # Recalculate descriptors and ensure correctness
        if ensure_correctness:
            smile_features = pca.transform(scaler.transform(pd.DataFrame({k: [v] for k, v in rdkit.Chem.Descriptors.CalcMolDescriptors(rdkit.Chem.MolFromSmiles(smile)).items()})[scaler.get_feature_names_out()]))
            assert np.allclose(smile_features[0], mol), "Features calculated from a smile string differ from features in the array"

    # Sample from each cluster    
    avg_len = np.mean([len(v) for v in cluster_to_mols.values()])
    cluster_to_samples = {}
    extra_mols = (100 - len(cluster_to_mols))*10
    left_to_sample = n_clusters*n_samples
    cluster_to_len = {cluster:len(mols) for cluster, mols in cluster_to_mols.items()}
    for i, (cluster, _) in enumerate(sorted(cluster_to_len.items(), key=lambda x: x[1], reverse=False)):
        smiles = cluster_to_mols[cluster]
        if extra_mols > 0:
            cur_extra = int(1+extra_mols/(len(cluster_to_mols) - i) * len(smiles)/avg_len)
            cur_samples = n_samples + cur_extra
            extra_mols -= cur_extra
        else:
            cur_samples = n_samples
        if cur_samples > left_to_sample:
            cur_samples = left_to_sample
        if len(smiles) > cur_samples:
            cluster_to_samples[cluster] = np.random.choice(smiles, cur_samples, replace=False)
            left_to_sample -= cur_samples
        else:
            cluster_to_samples[cluster] = smiles
            left_to_sample -= len(smiles)
            extra_mols += cur_samples - len(smiles)
    
    # Ensure the correct number of samples
    assert (n_sampled := sum(len(vals) for vals in cluster_to_samples.values())) == n_clusters * n_samples, f"Sampled {n_sampled} but were requested {n_clusters*n_samples}"

    # Save clusters and samples
    pickle.dump(cluster_to_mols, open(config_dict["clusters_save_path"], 'wb'))
    pickle.dump(cluster_to_samples, open(config_dict["samples_save_path"], 'wb'))

    # Create and save the final DataFrame
    keyToData = {}
    for cluster, mols in cluster_to_samples.items():
        for mol in mols:
            keyToData.setdefault('smiles', []).append(mol)
            keyToData.setdefault('cluster_id', []).append(cluster)
    pd.DataFrame(keyToData).to_csv(config["diffdock_save_path"])

    return cluster_to_samples
